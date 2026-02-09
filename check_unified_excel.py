"""통합 엑셀 파일 시트 목록 확인"""
from openpyxl import load_workbook
from pathlib import Path

# 가장 최근 통합 엑셀 파일 찾기
xlsx_files = sorted(Path('.').glob('all_results_*.xlsx'), key=lambda x: x.stat().st_mtime)

if not xlsx_files:
    print("No unified Excel files found")
else:
    xlsx_file = xlsx_files[-1]
    print(f"File: {xlsx_file.name}")
    print(f"Size: {xlsx_file.stat().st_size / 1024:.2f} KB")
    print()
    
    # 엑셀 파일 로드
    wb = load_workbook(xlsx_file)
    
    print(f"Total Sheets: {len(wb.sheetnames)}")
    print()
    print("Sheet List:")
    print("-" * 60)
    
    for idx, sheet_name in enumerate(wb.sheetnames, 1):
        ws = wb[sheet_name]
        print(f"{idx:2d}. {sheet_name:<40s} ({ws.max_row} rows)")
    
    print("-" * 60)
    print()
    print(f"SUCCESS: {xlsx_file.name} contains {len(wb.sheetnames)} sheets")
